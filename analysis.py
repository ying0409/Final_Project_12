import requests
from requests_html import HTMLSession
from bs4 import BeautifulSoup
from ckiptagger import data_utils
from ckiptagger import WS, POS, NER
from ckiptagger import construct_dictionary
import pandas as pd
import os
import openpyxl

class Word_Count_Analysis():
    def __init__(self,query_number):
        self.query_number = query_number
    def get_source(self,url):
        try:
            session = HTMLSession()
            response = session.get(url)
            return response
        except requests.exceptions.RequestException as e:
            print(e)
    # 網頁解析器
    def html_parser(self,htmlText):
        soup = BeautifulSoup(htmlText, 'html.parser')
        return soup
    def html_getText(self,soup):
        orignal_text = ''
        for el in soup.find_all('p'):
            orignal_text += ''.join(el.find_all(text=True))
        return orignal_text
    def read_excel_result(self,file):
        df = pd.read_excel(file)
        return df
        # book = openpyxl.load_workbook(os.path.join(file, 'result.xlsx'))
        # sheet = book['Sheet1']
        # nor = sheet.max_row
        # nol = sheet.max_column
        # head = [row for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True)][0]
        # datas = []
        # for row in sheet.iter_rows(min_row=2, max_row=nor + 1, values_only=True):
            # data = dict(zip(head, row))
            # datas.append(data)
        # return datas
    def word_count(self, texts):
        counts = dict()
        for text in texts:
            #stop_words = set(stopwords.words('english'))
            #words = word_tokenize(text)
            words = self.ckip_tokenize(text)
            for word in words[0]:
                #if word not in stop_words:
                if word in counts:
                    counts[word] += 1
                else:
                    counts[word] = 1
        return counts
    def ckip_tokenize(self, txts):
        #Add keywords to dictionary
        keyword_for_CKIP = {"台積電":1, "ASML":1, "Applied Materials":1, "SUMCO":1}
        dictionary_for_CKIP = construct_dictionary(keyword_for_CKIP)
        
        ws = WS("./data")
        words = ws([txts],coerce_dictionary=dictionary_for_CKIP)
        return words
    def get_wordcount_json(self, whitelists, index, dict_data, all_text, time):
        index = index%self.query_number
        data_array = []
        for i in whitelists[index]:
            if i in dict_data:
                if i=="台積電":
                    json_data = {
                        'Date' : time,
                        'Company' : i , 
                        'Count' : dict_data[i]
                    }
                else:
                    json_data = {
                        'Date' : time,
                        'Company' : i , 
                        'Count' : dict_data[i],
                        'Text' : all_text
                    }
                data_array.append(json_data)
        return data_array
    def jsonarray_toexcel(self,data_array):
        df = pd.DataFrame(data=data_array)
        df.to_excel('word_count.xlsx' , index=False)
        return

if __name__ == "__main__":
    path = "D:\Final_Project_12\\result.xlsx"
    analysis = Word_Count_Analysis(3)
    datas = analysis.read_excel_result(path)
    print(datas)
    print(datas["Date"].iloc[0])
    
    whitelists = [['台積電','ASML'],['台積電','Applied Materials'],['台積電','SUMCO']]
    count_array = []
    for index,links in enumerate(datas["Link"]):
        print(links)
        links = links.replace("['","")
        links = links.replace("']","")
        all_link = links.split("', '")
        print(all_link)
        all_text = []
        for link in all_link:
            print(link)
            Target_URL = link
            response = analysis.get_source(Target_URL)
            if response!=None:
                soup = analysis.html_parser(response.text)
                orignal_text = analysis.html_getText(soup)
                all_text.append(orignal_text)
            
        result_wordcount = analysis.word_count(all_text)
        print(result_wordcount)
        end_result = analysis.get_wordcount_json(whitelists, index, result_wordcount, all_text, datas["Date"].iloc[index])
        print(end_result)
        if len(end_result)!=0:
            if len(count_array)!=0:
                for result in end_result:
                    same = False
                    for count in count_array:
                        print(result,count)
                        if result['Date']==count['Date'] and result['Company']==count['Company']:
                            count['Count']+=result['Count']
                            same = True
                            break
                    if same==False:
                        count_array.append(result)
                        
            else:
                for result in end_result:
                    count_array.append(result)
        print(count_array)
    analysis.jsonarray_toexcel(count_array)
    print('Excel is OK')